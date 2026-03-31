#!/usr/bin/env python3
"""
parse_cases.py - 解析测试用例文件（Excel/CSV/JSON）
用法: python parse_cases.py <file_path> [--output JSON_PATH]
"""

import sys
import json
import csv
import argparse
from pathlib import Path

def parse_json(path: Path) -> list:
    with open(path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    if isinstance(data, list):
        return data
    return data.get('test_cases', data.get('cases', []))

def parse_csv(path: Path) -> list:
    cases = []
    with open(path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            cases.append({k.strip(): v.strip() if v else '' for k, v in row.items()})
    return cases

def parse_excel(path: Path) -> list:
    try:
        import openpyxl
    except ImportError:
        print("ERROR: 需要 openpyxl 库来读取 Excel 文件")
        print("安装方法: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else f'col_{i}'
              for i, cell in enumerate(ws[1])]

    cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            case = {}
            for i, (h, v) in enumerate(zip(headers, row)):
                case[h] = str(v).strip() if v is not None else ''
            cases.append(case)
    return cases

def parse_cases(file_path: str, output_path: str = None) -> list:
    path = Path(file_path)
    suffix = path.suffix.lower()

    if not path.exists():
        print(f"ERROR: 文件不存在: {file_path}")
        sys.exit(1)

    if suffix == '.json':
        cases = parse_json(path)
    elif suffix == '.csv':
        cases = parse_csv(path)
    elif suffix in ['.xlsx', '.xls']:
        cases = parse_excel(path)
    else:
        print(f"ERROR: 不支持的格式: {suffix}")
        print("支持的格式: .json, .csv, .xlsx, .xls")
        sys.exit(1)

    print(f"✅ 成功解析 {len(cases)} 个测试用例")

    # 打印前3个作为预览
    print("\n📋 前3个用例预览:")
    for i, case in enumerate(cases[:3]):
        print(f"  {i+1}. {case.get('name', '未命名')}")
        print(f"     type={case.get('type','N/A')} selector={case.get('selector','N/A')}")

    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(cases, f, ensure_ascii=False, indent=2)
        print(f"\n📝 已保存到: {output_path}")

    return cases

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='解析测试用例文件')
    parser.add_argument('file', help='测试用例文件路径 (.json/.csv/.xlsx/.xls)')
    parser.add_argument('--output', '-o', help='输出 JSON 路径（可选）')
    args = parser.parse_args()

    parse_cases(args.file, args.output)
