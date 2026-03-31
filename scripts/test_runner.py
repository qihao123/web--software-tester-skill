#!/usr/bin/env python3
"""
test_runner.py - Web 软件测试执行器
用法: python test_runner.py --config CONFIG_JSON --cases CASES_JSON [--output-dir DIR]
"""

import sys
import json
import csv
import argparse
import re
from pathlib import Path
from datetime import datetime

def parse_test_cases(cases_path: str) -> list:
    """解析测试用例文件，支持 JSON/CSV/Excel"""
    path = Path(cases_path)
    suffix = path.suffix.lower()

    if suffix == '.json':
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if isinstance(data, list):
                return data
            return data.get('test_cases', [])
    elif suffix == '.csv':
        cases = []
        with open(path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                cases.append(row)
        return cases
    elif suffix in ['.xlsx', '.xls']:
        # 延迟导入，优先用 openpyxl
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            cases = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    case = dict(zip(headers, row))
                    cases.append(case)
            return cases
        except ImportError:
            print("⚠️ 需要安装 openpyxl 才能读取 Excel 文件: pip install openpyxl")
            return []
    else:
        print(f"⚠️ 不支持的测试用例格式: {suffix}")
        return []

def run_tests(config: dict, cases: list, output_dir: str = '/tmp/test_results') -> dict:
    """
    执行测试的主逻辑
    config: {
        "url": str,
        "username": str,
        "password": str,
        "auto_explore": bool,
        "tests": [{"name": str, "type": str, "action": str, "selector": str, "value": str}]
    }
    """
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    results = {
        'config': config,
        'timestamp': timestamp,
        'total': 0,
        'passed': 0,
        'failed': 0,
        'skipped': 0,
        'tests': []
    }

    if not cases and not config.get('tests'):
        results['tests'].append({
            'name': '无测试用例',
            'status': 'skipped',
            'reason': '未提供测试用例',
            'timestamp': timestamp
        })
        results['skipped'] = 1
        return results

    # 执行测试（具体执行由 AI Agent 通过 browser 工具完成）
    results['message'] = '请使用 browser 工具执行测试，此脚本负责数据准备和报告生成'

    return results

def generate_report(results: dict, output_format: str = 'markdown', output_path: str = None):
    """生成测试报告"""
    timestamp = results.get('timestamp', datetime.now().strftime('%Y%m%d_%H%M%S'))

    if output_format == 'markdown':
        content = generate_markdown(results)
        ext = 'md'
    elif output_format == 'html':
        content = generate_html(results)
        ext = 'html'
    else:
        content = generate_markdown(results)
        ext = 'md'

    if not output_path:
        output_path = f'/tmp/test_report_{timestamp}.{ext}'

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(content)

    return output_path, content

def generate_markdown(results: dict) -> str:
    """生成 Markdown 格式报告"""
    total = results.get('total', 0)
    passed = results.get('passed', 0)
    failed = results.get('failed', 0)
    skipped = results.get('skipped', 0)
    pass_rate = (passed / total * 100) if total > 0 else 0

    md = f"""# 🧪 Web 软件测试报告

## 📊 测试概览

| 指标 | 数值 |
|------|------|
| 总用例数 | {total} |
| 通过 | ✅ {passed} |
| 失败 | ❌ {failed} |
| 跳过 | ⏭️ {skipped} |
| 通过率 | {pass_rate:.1f}% |

## 🕐 测试信息

- **测试时间**: {results.get('timestamp', 'N/A')}
- **目标URL**: {results.get('config', {}).get('url', 'N/A')}
- **测试模式**: {'自动探索' if results.get('config', {}).get('auto_explore') else '功能测试'}

## 📋 测试详情

"""

    for i, test in enumerate(results.get('tests', []), 1):
        status_icon = {'passed': '✅', 'failed': '❌', 'skipped': '⏭️'}.get(test.get('status'), '❓')
        md += f"""### {i}. {test.get('name', '未命名测试')} {status_icon}

| 项目 | 内容 |
|------|------|
| 状态 | {test.get('status', 'unknown').upper()} |
| 类型 | {test.get('type', 'N/A')} |
| 描述 | {test.get('description', 'N/A')} |

"""
        if test.get('error'):
            md += f"**❌ 错误信息**: {test.get('error')}\n\n"
        if test.get('screenshot'):
            md += f"**📸 截图**: `../{test.get('screenshot')}`\n\n"
        if test.get('response_time'):
            md += f"**⏱️ 响应时间**: {test.get('response_time')}ms\n\n"
        if test.get('bug_description'):
            md += f"**🐛 Bug 描述**: {test.get('bug_description')}\n\n"
        if test.get('steps'):
            md += "**测试步骤**:\n"
            for step in test.get('steps', []):
                md += f"- {step}\n"
            md += "\n"

    md += f"""## 💡 建议

"""
    if failed > 0:
        md += f"- 请优先修复 {failed} 个失败的测试用例\n"
    if pass_rate < 80:
        md += "- 通过率低于 80%，建议进行全面回归测试\n"
    else:
        md += "- 测试通过率良好，系统稳定性正常\n"

    return md

def generate_html(results: dict) -> str:
    """生成 HTML 格式报告（可打印为 PDF）"""
    pass_rate = (results.get('passed', 0) / results.get('total', 1) * 100)

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>Web 测试报告</title>
<style>
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; max-width: 900px; margin: 40px auto; padding: 20px; color: #333; }}
  h1 {{ color: #1a1a2e; }}
  .summary {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; margin: 24px 0; }}
  .card {{ background: #f8f9fa; border-radius: 12px; padding: 20px; text-align: center; }}
  .card.passed {{ background: #d4edda; }} .card.failed {{ background: #f8d7da; }}
  .card .num {{ font-size: 32px; font-weight: bold; }}
  .card .label {{ color: #666; margin-top: 4px; }}
  table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
  th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #eee; }}
  th {{ background: #f8f9fa; font-weight: 600; }}
  .passed {{ color: #28a745; }} .failed {{ color: #dc3545; }}
  .skipped {{ color: #ffc107; }}
  .bug {{ background: #fff3cd; padding: 12px; border-radius: 8px; margin: 8px 0; }}
  .screenshot {{ max-width: 100%; border: 1px solid #ddd; border-radius: 8px; margin: 8px 0; }}
</style>
</head>
<body>
<h1>🧪 Web 软件测试报告</h1>
<div class="summary">
  <div class="card"><div class="num">{results.get('total', 0)}</div><div class="label">总用例</div></div>
  <div class="card passed"><div class="num">{results.get('passed', 0)}</div><div class="label">通过</div></div>
  <div class="card failed"><div class="num">{results.get('failed', 0)}</div><div class="label">失败</div></div>
  <div class="card"><div class="num">{pass_rate:.1f}%</div><div class="label">通过率</div></div>
</div>
<p><strong>测试时间:</strong> {results.get('timestamp', 'N/A')} &nbsp;|&nbsp; <strong>URL:</strong> {results.get('config', {}).get('url', 'N/A')}</p>
<table>
<tr><th>#</th><th>用例名称</th><th>类型</th><th>状态</th><th>响应时间</th></tr>
"""
    for i, test in enumerate(results.get('tests', []), 1):
        status = test.get('status', 'unknown')
        html += f"""<tr>
  <td>{i}</td>
  <td>{test.get('name', 'N/A')}</td>
  <td>{test.get('type', 'N/A')}</td>
  <td class="{status}">{status.upper()}</td>
  <td>{test.get('response_time', 'N/A')} ms</td>
</tr>
"""
        if test.get('bug_description'):
            html += f'<tr><td colspan="5"><div class="bug">🐛 <strong>Bug:</strong> {test.get("bug_description")}</div></td></tr>'

    html += """</table></body></html>"""
    return html

def main():
    parser = argparse.ArgumentParser(description='Web 软件测试执行器')
    parser.add_argument('--config', help='测试配置 JSON 字符串或文件路径')
    parser.add_argument('--cases', help='测试用例 JSON/CSV/Excel 文件路径')
    parser.add_argument('--output-dir', default='/tmp/test_results', help='输出目录')
    parser.add_argument('--format', choices=['markdown', 'html'], default='markdown', help='报告格式')
    parser.add_argument('--report-output', help='报告输出路径')
    args = parser.parse_args()

    # 解析配置
    if args.config:
        if Path(args.config).exists():
            with open(args.config, 'r', encoding='utf-8') as f:
                config = json.load(f)
        else:
            config = json.loads(args.config)
    else:
        config = {}

    # 解析用例
    cases = []
    if args.cases and Path(args.cases).exists():
        cases = parse_test_cases(args.cases)
        print(f"📋 已加载 {len(cases)} 个测试用例")

    results = run_tests(config, cases, args.output_dir)
    report_path, _ = generate_report(results, args.format, args.report_output)
    print(f"✅ 报告已生成: {report_path}")

if __name__ == '__main__':
    main()
